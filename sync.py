#!/usr/bin/env python3
"""Scrape the Fraternus chapter roster and build a Flocknote mail-list import.

One row per distinct email address (a household's youth members carry the
account holder's address, so they add nothing to a mail list). Members with no
email are dropped. Emails seen on previous runs are tracked so each run can
also emit a "new since last time" file to upload.

Usage:
    python3 sync.py              # scrape browser, write import files
    python3 sync.py --dry-run    # scrape and report, write nothing
    python3 sync.py --reset      # forget history, treat everyone as new

Requires Chromium running with --remote-debugging-port=9222, logged in.
"""
import argparse, csv, datetime, json, pathlib, re, subprocess, sys

HERE = pathlib.Path(__file__).parent.resolve()

# Flocknote's own import template. Optional: without it we still write the CSV,
# which is all the importer actually needs. Override with --template or
# FLOCKNOTE_TEMPLATE. Download it from Flocknote's import page.
TEMPLATE_NAME = "examplecompletespreadsheet.xlsx"
TEMPLATE_SEARCH = [
    HERE / TEMPLATE_NAME,
    pathlib.Path.home() / "Downloads" / TEMPLATE_NAME,
]
STATE = HERE / "seen_emails.json"
RAW = HERE / "roster_raw.json"
PORTAL = "portal.fraternus.org"   # which browser tab to drive
OUT_XLSX = HERE / "flocknote_import.xlsx"
OUT_CSV = HERE / "flocknote_import.csv"
OUT_NEW = HERE / "flocknote_new.csv"

COLS = ["First Name", "Last Name", "Email", "Phone", "Birthday", "Home Street",
        "Home City", "Home State", "Home Zip", "Shirt Size", "Class Year"]
HEADER_ROW, FIRST_DATA_ROW = 2, 3


def scrape():
    """Drive the logged-in browser and return the roster payload."""
    try:
        proc = subprocess.run(
            ["node", str(HERE / "cdp.js"), "eval", str(HERE / "scrape.js"), PORTAL],
            capture_output=True, text=True, timeout=600, cwd=HERE)
    except FileNotFoundError:
        sys.exit("Node.js is required but 'node' was not found on PATH.")
    except subprocess.TimeoutExpired:
        sys.exit("Timed out driving the browser. Is the roster page loaded?")
    if proc.returncode != 0:
        sys.exit("Scrape failed:\n" + (proc.stderr or proc.stdout).strip())
    try:
        data = json.loads(proc.stdout.strip())
    except json.JSONDecodeError:
        sys.exit("Unexpected output from browser:\n" + proc.stdout[:500])
    if data.get("error"):
        hint = {
            "not-logged-in": f"Log in to {PORTAL} in the debug Chromium first.",
            "no-my-chapter-nav": "Could not find the MY CHAPTER nav button.",
            "no-members-found": "Roster list was empty - did the page finish loading?",
        }.get(data["error"], "")
        sys.exit(f"Scrape error: {data['error']}. {hint}")
    return data


def fmt_phone(p):
    d = re.sub(r"\D", "", p or "")
    if len(d) == 11 and d.startswith("1"):
        d = d[1:]
    return f"{d[0:3]}-{d[3:6]}-{d[6:]}" if len(d) == 10 else (p or "")


def split_name(full):
    parts = full.split()
    return (parts[0], " ".join(parts[1:])) if len(parts) > 1 else (full, "")


def dedupe(members):
    """One entry per email. Prefer an ADULT as the contact name."""
    chosen, extras = {}, {}
    for m in members:
        email = (m.get("email") or "").strip().lower()
        if not email:
            continue
        extras.setdefault(email, []).append(m)
        cur = chosen.get(email)
        if cur is None or (cur["kind"] != "ADULT" and m["kind"] == "ADULT"):
            chosen[email] = m
    return chosen, extras


def build_rows(chosen):
    rows = []
    for email, m in sorted(chosen.items(), key=lambda kv: kv[1]["name"].split()[-1].lower()):
        first, last = split_name(m["name"])
        row = dict.fromkeys(COLS, "")
        row["First Name"], row["Last Name"] = first, last
        row["Email"], row["Phone"] = m["email"], fmt_phone(m["phone"])
        rows.append(row)
    return rows


def find_template(explicit=None):
    for cand in ([pathlib.Path(explicit)] if explicit else []) + TEMPLATE_SEARCH:
        if cand.exists():
            return cand
    return None


def write_xlsx(rows, explicit=None):
    template = find_template(explicit)
    if template is None:
        print(f"  - no {TEMPLATE_NAME} found; wrote CSV only (that's fine for import)")
        return False
    try:
        import openpyxl
    except ImportError:
        print("  - openpyxl not installed; wrote CSV only (pip install openpyxl for .xlsx)")
        return False
    wb = openpyxl.load_workbook(template)
    ws = wb.active
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(1, len(COLS) + 1):
            ws.cell(row=r, column=c).value = None
    for i, m in enumerate(rows):
        for c, col in enumerate(COLS, start=1):
            ws.cell(row=FIRST_DATA_ROW + i, column=c).value = m[col] or None
    wb.save(OUT_XLSX)
    return True


def write_csv(path, rows):
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        w.writerows(rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="report only, write nothing")
    ap.add_argument("--reset", action="store_true", help="forget history; everyone is new")
    ap.add_argument("--template", help=f"path to Flocknote's {TEMPLATE_NAME}")
    args = ap.parse_args()

    data = scrape()
    members = data["members"]
    chosen, extras = dedupe(members)
    rows = build_rows(chosen)

    no_email = [m["name"] for m in members if not m.get("email")]
    seen = set() if args.reset else set(json.loads(STATE.read_text())["emails"]) if STATE.exists() else set()
    new_rows = [r for r in rows if r["Email"].lower() not in seen]

    print(f"chapter        : {data.get('chapter', '?')}")
    print(f"scraped        : {len(members)} members")
    print(f"mailable       : {len(rows)} unique emails")
    print(f"no email       : {len(no_email)}" + (f" ({', '.join(no_email)})" if no_email else ""))
    print(f"new since last : {len(new_rows)}" if STATE.exists() and not args.reset
          else f"new since last : {len(new_rows)} (no history yet)")
    for r in new_rows:
        print(f"    + {r['First Name']} {r['Last Name']} <{r['Email']}>")

    # Contacts whose only roster name is a youth - the address is really a parent's.
    odd = [(e, ms[0]) for e, ms in extras.items()
           if not any(m["kind"] == "ADULT" for m in ms)]
    if odd:
        print(f"\n  ! {len(odd)} email(s) attached only to a YOUTH record - the address"
              f"\n    belongs to a parent who is not a chapter member. Check the name:")
        for e, m in odd:
            print(f"      {e}  -> listed as {m['name']}")

    if args.dry_run:
        print("\n(dry run - nothing written)")
        return

    RAW.write_text(json.dumps(data, indent=2))
    write_csv(OUT_CSV, rows)
    wrote_xlsx = write_xlsx(rows, args.template)
    if new_rows:
        write_csv(OUT_NEW, new_rows)

    STATE.write_text(json.dumps({
        "updated": datetime.datetime.now().isoformat(timespec="seconds"),
        "emails": sorted(set(seen) | {r["Email"].lower() for r in rows}),
    }, indent=2))

    print(f"\nwrote {OUT_CSV.name}" + (f", {OUT_XLSX.name}" if wrote_xlsx else "")
          + f" ({len(rows)} contacts)")
    if new_rows:
        print(f"wrote {OUT_NEW.name} ({len(new_rows)} new) <- upload this one")
    else:
        print("no new contacts since last run - nothing to upload")


if __name__ == "__main__":
    main()
